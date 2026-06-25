import streamlit as st
import pandas as pd
import geopandas as gpd
import requests
from shapely.geometry import LineString
import time
import openrouteservice
import os
import json
import tempfile
import io
import folium
from streamlit_folium import st_folium
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import gdown
import zipfile

GDRIVE_ID = "1eeobEZW0gTqdXULMpNS5a7EHKtNAfE8y"

def download_shapefile_if_needed():
    if not os.path.exists("shapefile/tolwegen.shp"):
        url = f"https://drive.google.com/uc?id={GDRIVE_ID}"
        gdown.download(url, "shapefile.zip", quiet=False)
        with zipfile.ZipFile("shapefile.zip", "r") as z:
            z.extractall(".")
        os.remove("shapefile.zip")

download_shapefile_if_needed()

st.set_page_config(page_title="VWH Tool — Kamps Transport", layout="wide", page_icon="🚛")

# ─── Constanten ───────────────────────────────────────────────────────────────
GEOCODE_CACHE_FILE = "geocode_cache.json"
ROUTE_CACHE_FILE   = "route_cache.json"
POSTCODE_FILE      = "NL3.xlsx"
DE_POSTCODE_FILE   = "DE.xlsx"
BE_POSTCODE_FILE   = "BE.xlsx"
SHAPEFILE_TOL      = "shapefile/tolwegen.shp"
SHAPEFILE_COUNTRIES = "shapefile/ne_10m_admin_0_countries_deu.shp"
BLEISWIJK_PLAATS   = "Bleiswijk"
BLEISWIJK_POSTCODE = "2665MZ"
MAX_GEWICHT        = 24000

# ─── Cache ────────────────────────────────────────────────────────────────────
def load_cache(filepath):
    if not os.path.exists(filepath): return {}
    try:
        with open(filepath, "r") as f: return json.load(f)
    except json.JSONDecodeError: return {}

def save_cache(data, filepath):
    dir_ = os.path.dirname(os.path.abspath(filepath))
    with tempfile.NamedTemporaryFile("w", dir=dir_, delete=False, suffix=".tmp") as tmp:
        json.dump(data, tmp); tmp_path = tmp.name
    os.replace(tmp_path, filepath)

# ─── Resources (gecached) ─────────────────────────────────────────────────────
@st.cache_resource
def load_resources():
    roads    = gpd.read_file(SHAPEFILE_TOL)
    tolwegen = roads[roads["HEFNETWERK"] == "Ja"].to_crs(epsg=28992)
    tol_sindex = tolwegen.sindex
    countries = gpd.read_file(SHAPEFILE_COUNTRIES).to_crs(epsg=28992)
    nl_geom = countries[countries["NAME"] == "Netherlands"].geometry.iloc[0]
    de_geom = countries[countries["NAME"] == "Germany"].geometry.iloc[0]
    be_geom = countries[countries["NAME"] == "Belgium"].geometry.iloc[0]
    postcode_df = pd.read_excel(POSTCODE_FILE)
    postcode_df["postcode"] = postcode_df["Postcode"].astype(str).str[:4]
    postcode_df["plaats"]   = postcode_df["Plaats"].str.lower()
    de_df = pd.read_excel(DE_POSTCODE_FILE)
    de_df["postcode"] = de_df["Postcode"].astype(str)
    de_df["plaats"]   = de_df["Plaats"].str.lower()
    be_df = pd.read_excel(BE_POSTCODE_FILE)
    be_df["postcode"] = be_df["Postcode"].astype(str)
    be_df["plaats"]   = be_df["Plaats"].str.lower()
    return tolwegen, tol_sindex, nl_geom, de_geom, be_geom, postcode_df, de_df, be_df

# ─── Postcode helpers ─────────────────────────────────────────────────────────
def is_nl_postcode(p):
    if not p: return False
    p = str(p).replace(" ","").upper()
    return (len(p)==4 and p.isdigit()) or (len(p)==6 and p[:4].isdigit() and p[4:].isalpha())

def is_de_postcode(p):
    if not p: return False
    p = str(p).replace(" ","")
    return p.isdigit() and len(p)==5

def is_be_postcode(p):
    if not p: return False
    p = str(p).replace(" ","")
    return p.isdigit() and len(p)==4

def lookup_postcode(postcode, plaats, df):
    matches = df[df["postcode"] == postcode]
    if len(matches)==0: return None
    for _, row in matches.iterrows():
        if str(plaats).lower() in row["plaats"]: return [row["Lon"], row["Lat"]]
    return [matches.iloc[0]["Lon"], matches.iloc[0]["Lat"]]

# ─── Geocode ──────────────────────────────────────────────────────────────────
def geocode(place, postcode, api_key, geocode_cache, postcode_df, de_df, be_df, landcode=None):
    place_clean    = str(place).strip().lower()
    postcode_clean = "" if pd.isna(postcode) else str(postcode).strip()
    land = str(landcode).strip().upper() if landcode and not pd.isna(landcode) else None
    cache_key = f"{place_clean}_{postcode_clean}_{land or ''}"
    if cache_key in geocode_cache: return geocode_cache[cache_key]
    coords = None
    if land is None or land == "NL":
        if is_nl_postcode(postcode_clean):
            coords = lookup_postcode(postcode_clean.replace(" ","")[:4], place_clean, postcode_df)
    if not coords and (land is None or land == "DE"):
        if is_de_postcode(postcode_clean):
            coords = lookup_postcode(postcode_clean.replace(" ",""), place_clean, de_df)
    if not coords and (land is None or land == "BE"):
        if is_be_postcode(postcode_clean):
            coords = lookup_postcode(postcode_clean.replace(" ",""), place_clean, be_df)
    if not coords:
        boundary = land if land in ("NL","DE","BE") else "NL,DE,BE"
        r = requests.get("https://api.openrouteservice.org/geocode/search",
                         headers={"Authorization": api_key},
                         params={"text": f"{postcode_clean} {place_clean}".strip(),
                                 "boundary.country": boundary, "size": 5})
        if r.status_code == 200:
            data = r.json()
            if data.get("features"):
                lon, lat = data["features"][0]["geometry"]["coordinates"]
                coords = [lon, lat]
    if coords:
        geocode_cache[cache_key] = coords
        save_cache(geocode_cache, GEOCODE_CACHE_FILE)
    return coords

# ─── Berekeningen ─────────────────────────────────────────────────────────────
def calculate_country_km(coords, geom):
    if len(coords) < 2: return 0
    line = LineString(coords)
    gdf  = gpd.GeoDataFrame(geometry=[line], crs="EPSG:4326").to_crs(epsg=28992)
    return gdf.geometry.iloc[0].intersection(geom).length / 1000

def calculate_tol_km(coords, tolwegen, tol_sindex):
    if len(coords) < 2: return 0, []
    line    = LineString(coords)
    gdf     = gpd.GeoDataFrame(geometry=[line], crs="EPSG:4326").to_crs(tolwegen.crs)
    line_rd = gdf.geometry.iloc[0]
    idx     = list(tol_sindex.intersection(line_rd.bounds))
    if not idx:
        line_wgs = gpd.GeoDataFrame(geometry=[line_rd], crs=tolwegen.crs).to_crs("EPSG:4326").geometry.iloc[0]
        return 0, [(False, list(line_wgs.coords))]
    kandidaten = tolwegen.iloc[idx]
    route_buf  = line_rd.buffer(20)
    tol_union  = kandidaten[kandidaten.intersects(route_buf)].geometry.union_all()
    if tol_union is None or tol_union.is_empty:
        line_wgs = gpd.GeoDataFrame(geometry=[line_rd], crs=tolwegen.crs).to_crs("EPSG:4326").geometry.iloc[0]
        return 0, [(False, list(line_wgs.coords))]
    tol_buf  = tol_union.buffer(20)
    tol_deel = line_rd.intersection(tol_buf)
    tol_m    = tol_deel.length if not tol_deel.is_empty else 0
    dist = 0; segmenten = []
    while dist < line_rd.length:
        seg    = LineString([line_rd.interpolate(dist), line_rd.interpolate(min(dist+100, line_rd.length))])
        is_tol = seg.intersects(tol_buf)
        seg_wgs = gpd.GeoDataFrame(geometry=[seg], crs=tolwegen.crs).to_crs("EPSG:4326").geometry.iloc[0]
        segmenten.append((is_tol, list(seg_wgs.coords)))
        dist += 100
    return tol_m / 1000, segmenten

def make_route_key(start, end):
    return f"{round(start[0],5)},{round(start[1],5)}_{round(end[0],5)},{round(end[1],5)}"

def get_route(start, end, route_cache, api_key, tolwegen, tol_sindex, nl_geom, de_geom, be_geom, route_sleep):
    key = make_route_key(start, end)
    if key in route_cache:
        data = route_cache[key]
        _, segmenten = calculate_tol_km(data["coords"], tolwegen, tol_sindex)
        return (data["NL"], data["DE"], data["BE"], data["TOL"], data["coords"], segmenten)
    def req(profile):
        r = requests.post(f"https://api.openrouteservice.org/v2/directions/{profile}",
                          json={"coordinates": [start, end]},
                          headers={"Authorization": api_key, "Content-Type": "application/json"})
        if r.status_code != 200: return None
        d = r.json()
        if not d.get("routes"): return None
        return openrouteservice.convert.decode_polyline(d["routes"][0]["geometry"])["coordinates"]
    time.sleep(route_sleep)
    coords = req("driving-hgv") or req("driving-car")
    if not coords: return None
    nl  = calculate_country_km(coords, nl_geom)
    de  = calculate_country_km(coords, de_geom)
    be  = calculate_country_km(coords, be_geom)
    tol, segmenten = calculate_tol_km(coords, tolwegen, tol_sindex)
    route_cache[key] = {"NL": nl, "DE": de, "BE": be, "TOL": tol, "coords": coords}
    if len(route_cache) % 50 == 0:
        save_cache(route_cache, ROUTE_CACHE_FILE)
    return (nl, de, be, tol, coords, segmenten)

# ─── Kaart ────────────────────────────────────────────────────────────────────
def maak_kaart(label, segmenten):
    alle_coords = [pt for _, seg in segmenten for pt in seg]
    if not alle_coords: return None
    mid_lat = sum(c[1] for c in alle_coords) / len(alle_coords)
    mid_lon = sum(c[0] for c in alle_coords) / len(alle_coords)
    m = folium.Map(location=[mid_lat, mid_lon], zoom_start=8, tiles="CartoDB positron")
    for is_tol, seg_coords in segmenten:
        latlons = [(lat, lon) for lon, lat in seg_coords]
        if len(latlons) >= 2:
            folium.PolyLine(latlons, color="#E8290B" if is_tol else "#2563EB",
                            weight=5 if is_tol else 3, opacity=0.85,
                            tooltip="Tolweg" if is_tol else "Geen tol").add_to(m)
    legenda = f"""<div style="position:fixed;bottom:30px;left:30px;z-index:9999;
        background:white;padding:10px 14px;border-radius:8px;
        box-shadow:0 2px 8px rgba(0,0,0,0.2);font-family:Arial;font-size:13px;">
        <b>{label}</b><br>
        <span style="color:#E8290B;">&#9644;</span> Tolweg<br>
        <span style="color:#2563EB;">&#9644;</span> Geen tol</div>"""
    m.get_root().html.add_child(folium.Element(legenda))
    return m

# ─── Excel output ─────────────────────────────────────────────────────────────
def maak_excel(output_results, klant_rows):
    NAVY="1B2A4A"; GRAY_HD="3D5278"; TOT_BG="D6E4F7"; TOT_FG="0C2D6B"
    DEPOT_BG="F4F4F4"; DEPOT_FG="888888"; WHITE="FFFFFF"; ROW_ALT="F8FBFF"
    BORDER_C="CCDAEE"; GREEN_FG="0C4A1E"
    def mk_side(c=BORDER_C, s="thin"): return Side(border_style=s, color=c)
    def mk_border(c=BORDER_C, s="thin"):
        side = mk_side(c,s); return Border(left=side,right=side,top=side,bottom=side)
    def mk_fill(c): return PatternFill("solid", fgColor=c)
    KM_FMT="#,##0.00"; EUR_FMT="€\\ #,##0.00"; KG_FMT="#,##0"
    COLS=["Rit","Van","Naar","Klant","Opdracht nr","NL km","DE km","BE km","Tol km","Kosten €",
          "Km direct (Bleiswijk→stop)","Tol km direct (Bleiswijk→stop)","Kosten € (direct)",
          "Kosten toegerekend €","Km toegerekend","Gewicht (kg)"]
    HDRS=["Rit","Van","Naar","Klant","Opdracht nr","NL km","DE km","BE km","Tol km","Kosten €",
          "Km direct\n(Bleiswijk→stop)","Tol km direct\n(Bleiswijk→stop)","Kosten €\n(direct)",
          "Kosten\ntoegerekend €","Km\ntoegerekend","Gewicht (kg)"]
    WIDTHS=[6,18,22,32,13,9,9,9,9,11,18,18,13,18,14,13]
    KM_COLS={6,7,8,9,11,12,15}; EUR_COLS={10,13,14}; KG_COLS={16}
    wb = Workbook(); ws = wb.active; ws.title = "Ritten"
    for c,(label,width) in enumerate(zip(HDRS,WIDTHS),1):
        cell = ws.cell(row=1,column=c,value=label)
        cell.font=Font(name="Arial",bold=True,color=WHITE,size=10)
        cell.fill=mk_fill(GRAY_HD if c in {10,13,14} else NAVY)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=mk_border(color=NAVY)
        ws.column_dimensions[get_column_letter(c)].width=width
    ws.row_dimensions[1].height=34
    alt=0
    for row_data in output_results:
        is_totaal = row_data.get("Van")=="TOTAAL"
        is_depot  = row_data.get("Van")==BLEISWIJK_PLAATS and row_data.get("Naar")==BLEISWIJK_PLAATS
        if not is_totaal and not is_depot: alt+=1
        er = ws.max_row+1; ws.row_dimensions[er].height=16
        for c,col in enumerate(COLS,1):
            val=row_data.get(col,"")
            cell=ws.cell(row=er,column=c,value=val if val!="" else None)
            if is_totaal:
                cell.font=Font(name="Arial",bold=True,color=TOT_FG,size=10)
                cell.fill=mk_fill(TOT_BG); cell.border=mk_border(color=BORDER_C)
            elif is_depot:
                cell.font=Font(name="Arial",italic=True,color=DEPOT_FG,size=9)
                cell.fill=mk_fill(DEPOT_BG); cell.border=mk_border(color="DDDDDD",style="hair")
            else:
                cell.font=Font(name="Arial",size=10)
                cell.fill=mk_fill(ROW_ALT) if alt%2==0 else mk_fill(WHITE)
                cell.border=mk_border(color=BORDER_C,style="hair")
            if c in KM_COLS: cell.number_format=KM_FMT; cell.alignment=Alignment(horizontal="right",vertical="center")
            elif c in EUR_COLS:
                cell.number_format=EUR_FMT; cell.alignment=Alignment(horizontal="right",vertical="center")
                if c==14 and not is_totaal and not is_depot: cell.font=Font(name="Arial",size=10,color=GREEN_FG)
            elif c in KG_COLS: cell.number_format=KG_FMT; cell.alignment=Alignment(horizontal="right",vertical="center")
            elif c==1: cell.alignment=Alignment(horizontal="center",vertical="center")
            elif c==5: cell.alignment=Alignment(horizontal="center",vertical="center")
            elif c in {2,3,4}: cell.alignment=Alignment(horizontal="left",vertical="center")
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(COLS))}1"
    ws2=wb.create_sheet("Klant totalen")
    h2=["Klant","NL km","DE km","BE km","Tol km","Kosten €","Gem km per rit"]
    w2=[38,13,13,13,13,14,14]
    for c,(h,w) in enumerate(zip(h2,w2),1):
        cell=ws2.cell(row=1,column=c,value=h)
        cell.font=Font(name="Arial",bold=True,color=WHITE,size=10)
        cell.fill=mk_fill(NAVY); cell.alignment=Alignment(horizontal="center",vertical="center")
        cell.border=mk_border(color=NAVY); ws2.column_dimensions[get_column_letter(c)].width=w
    ws2.row_dimensions[1].height=24
    for r,krow in enumerate(klant_rows,1):
        er=r+1; ws2.row_dimensions[er].height=16
        vals=[krow["Klant"],krow["NL km"],krow["DE km"],krow["BE km"],
              krow["Tol km"],krow["Kosten €"],krow["Gem km per rit"]]
        for c,val in enumerate(vals,1):
            cell=ws2.cell(row=er,column=c,value=val)
            cell.fill=mk_fill(ROW_ALT) if r%2==0 else mk_fill(WHITE)
            cell.border=mk_border(color=BORDER_C,style="hair")
            cell.alignment=Alignment(horizontal="left" if c==1 else "right",vertical="center")
            if c in {2,3,4,5,7}: cell.number_format=KM_FMT; cell.font=Font(name="Arial",size=10)
            elif c==6: cell.number_format=EUR_FMT; cell.font=Font(name="Arial",bold=True,size=10,color=TOT_FG)
            else: cell.font=Font(name="Arial",size=10)
    ws2.freeze_panes="A2"; ws2.auto_filter.ref="A1:G1"
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ─── Hoofd berekening ─────────────────────────────────────────────────────────
def bereken(df, api_key, tarief, route_sleep, geocode_cache, route_cache,
            tolwegen, tol_sindex, nl_geom, de_geom, be_geom,
            postcode_df, de_df, be_df, voortgang_bar, status_tekst):

    plaatsen    = df.iloc[:,0].tolist()
    postcodes   = df.iloc[:,1].tolist()
    klanten     = df.iloc[:,2].tolist()
    ritten      = df.iloc[:,3].tolist()
    gewichten   = df.iloc[:,4].tolist()
    opdrachtnrs = df.iloc[:,5].tolist() if df.shape[1]>5 else [None]*len(plaatsen)
    landcodes   = df.iloc[:,6].tolist() if df.shape[1]>6 else [None]*len(plaatsen)

    results=[]; rit_totalen={"NL":0,"DE":0,"BE":0,"TOL":0,"KOSTEN":0}
    rit_segmenten_huidig=[]; rit_kaarten={}
    prev_place=prev_pc=prev_klant=prev_rit=prev_gewicht=prev_opdrachtnr=prev_land=None
    totaal=len(plaatsen)

    bleiswijk_coords = geocode(BLEISWIJK_PLAATS, BLEISWIJK_POSTCODE, api_key,
                               geocode_cache, postcode_df, de_df, be_df, "NL")

    for i,(place,pc,klant,rit,gewicht,opdrachtnr,landcode) in enumerate(
            zip(plaatsen,postcodes,klanten,ritten,gewichten,opdrachtnrs,landcodes)):
        voortgang_bar.progress(i/totaal, text=f"Rij {i+1}/{totaal}")
        if pd.isna(place): continue
        place=str(place).strip()

        if prev_rit is not None and rit != prev_rit:
            results.append({"Rit":prev_rit,"Van":"TOTAAL","Naar":"","Klant":"","Opdracht nr":"",
                "NL km":round(rit_totalen["NL"],2),"DE km":round(rit_totalen["DE"],2),
                "BE km":round(rit_totalen["BE"],2),"Tol km":round(rit_totalen["TOL"],2),
                "Kosten €":round(rit_totalen["KOSTEN"],2),"Km direct (Bleiswijk→stop)":"",
                "Kosten € (direct)":"","_gewicht":None,"_is_totaal":True})
            rit_kaarten[prev_rit] = rit_segmenten_huidig[:]
            rit_totalen={"NL":0,"DE":0,"BE":0,"TOL":0,"KOSTEN":0}
            rit_segmenten_huidig=[]; prev_place=None

        if prev_place is None:
            prev_place=place;prev_pc=pc;prev_klant=klant
            prev_rit=rit;prev_gewicht=gewicht;prev_opdrachtnr=opdrachtnr;prev_land=landcode
            continue

        status_tekst.write(f"🔄 **{prev_place}** → **{place}**")
        start=geocode(prev_place,prev_pc,api_key,geocode_cache,postcode_df,de_df,be_df,prev_land)
        end  =geocode(place,pc,api_key,geocode_cache,postcode_df,de_df,be_df,landcode)

        route=None
        if start and end:
            key=make_route_key(start,end)
            if key not in route_cache: pass
            route=get_route(start,end,route_cache,api_key,tolwegen,tol_sindex,nl_geom,de_geom,be_geom,route_sleep)

        if not route:
            prev_place=place;prev_pc=pc;prev_klant=klant
            prev_rit=rit;prev_gewicht=gewicht;prev_opdrachtnr=opdrachtnr;prev_land=landcode
            continue

        nl,de,be,tol,coords,segmenten=route
        rit_segmenten_huidig.extend(segmenten)

        direct_tol=direct_kosten=direct_totaal_km=0
        if bleiswijk_coords and end:
            rd=get_route(bleiswijk_coords,end,route_cache,api_key,tolwegen,tol_sindex,nl_geom,de_geom,be_geom,route_sleep)
            if rd:
                nl_d,de_d,be_d,direct_tol,_,_=rd
                direct_kosten=direct_tol*tarief; direct_totaal_km=nl_d+de_d+be_d

        gewicht_val=None
        if not pd.isna(gewicht) and str(gewicht).strip()!="":
            try: gewicht_val=float(gewicht)
            except: pass

        try: opdr_str="" if pd.isna(opdrachtnr) else str(int(float(opdrachtnr))) if opdrachtnr else ""
        except: opdr_str=str(opdrachtnr) if opdrachtnr else ""

        results.append({"Rit":rit,"Van":prev_place,"Naar":place,"Klant":klant,"Opdracht nr":opdr_str,
            "NL km":round(nl,2),"DE km":round(de,2),"BE km":round(be,2),
            "Tol km":round(tol,2),"Kosten €":round(tol*tarief,2),
            "Km direct (Bleiswijk→stop)":round(direct_totaal_km,2),
            "Tol km direct (Bleiswijk→stop)":round(direct_tol,2),
            "Kosten € (direct)":round(direct_kosten,2),
            "_gewicht":gewicht_val,"_is_totaal":False})

        rit_totalen["NL"]+=nl;rit_totalen["DE"]+=de
        rit_totalen["BE"]+=be;rit_totalen["TOL"]+=tol
        rit_totalen["KOSTEN"]+=tol*tarief
        prev_place=place;prev_pc=pc;prev_klant=klant
        prev_rit=rit;prev_gewicht=gewicht;prev_opdrachtnr=opdrachtnr;prev_land=landcode

    if prev_rit is not None:
        results.append({"Rit":prev_rit,"Van":"TOTAAL","Naar":"","Klant":"","Opdracht nr":"",
            "NL km":round(rit_totalen["NL"],2),"DE km":round(rit_totalen["DE"],2),
            "BE km":round(rit_totalen["BE"],2),"Tol km":round(rit_totalen["TOL"],2),
            "Kosten €":round(rit_totalen["KOSTEN"],2),"Km direct (Bleiswijk→stop)":"",
            "Kosten € (direct)":"","_gewicht":None,"_is_totaal":True})
        rit_kaarten[prev_rit] = rit_segmenten_huidig[:]

    voortgang_bar.progress(1.0, text="COFRET berekenen...")

    # COFRET
    rit_werkelijke_kosten={}; rit_werkelijke_alle_km={}
    for row in results:
        if row.get("_is_totaal"):
            rv=row.get("Rit")
            if rv!="" and not pd.isna(rv):
                rit_werkelijke_kosten[rv]=row["Kosten €"]
                rit_werkelijke_alle_km[rv]=row["NL km"]+row["DE km"]+row["BE km"]

    opdr_gew={}; opdr_direct_km_som={}; opdr_direct_km_cnt={}; opdr_direct_eerste={}; gezien_toerek=set()
    for row in results:
        if row.get("_is_totaal"): continue
        opdr=str(row.get("Opdracht nr") or "")
        if not opdr: continue
        rit=row.get("Rit"); key=(rit,opdr)
        opdr_gew[key]=opdr_gew.get(key,0)+(row.get("_gewicht") or 0)
        if key not in gezien_toerek:
            gezien_toerek.add(key)
            km=row.get("Km direct (Bleiswijk→stop)",0) or 0
            opdr_direct_km_som[key]=km; opdr_direct_km_cnt[key]=1
            kosten=row.get("Kosten € (direct)",0) or 0
            opdr_direct_eerste[key]={"som":kosten,"cnt":1}

    opdr_direct_km={k:opdr_direct_km_som[k]/opdr_direct_km_cnt[k] for k in opdr_direct_km_som}
    opdr_direct={k:v["som"]/v["cnt"] for k,v in opdr_direct_eerste.items()}
    opdr_vwh={k:(opdr_direct.get(k,0))*(gew/MAX_GEWICHT) for k,gew in opdr_gew.items() if gew>0}
    rit_vwh_som={}
    for (rit,opdr),vwh in opdr_vwh.items(): rit_vwh_som[rit]=rit_vwh_som.get(rit,0)+vwh
    rit_direct_km_som={}
    for (rit,opdr),km in opdr_direct_km.items(): rit_direct_km_som[rit]=rit_direct_km_som.get(rit,0)+km
    opdr_km_toegerekend={}
    for (rit,opdr),dk in opdr_direct_km.items():
        dks=rit_direct_km_som.get(rit,0); wk=rit_werkelijke_alle_km.get(rit,0)
        if dk and dks: opdr_km_toegerekend[(rit,opdr)]=round((dk/dks)*wk,2)

    _gezien=set()
    for row in results:
        if row.get("_is_totaal"): row["Kosten toegerekend €"]=""; row["Km toegerekend"]=""; continue
        rit=row.get("Rit"); opdr=str(row.get("Opdracht nr") or "")
        if (rit,opdr) not in _gezien:
            _gezien.add((rit,opdr))
            vwh=opdr_vwh.get((rit,opdr),0); vs=rit_vwh_som.get(rit,0); wk=rit_werkelijke_kosten.get(rit,0)
            row["Kosten toegerekend €"]=round((vwh/vs)*wk,2) if vwh and vs else ""
            row["Km toegerekend"]=opdr_km_toegerekend.get((rit,opdr),"")
        else: row["Kosten toegerekend €"]=""; row["Km toegerekend"]=""

    rit_samenvatting={}
    for row in results:
        if row.get("Van")=="TOTAAL":
            rv=row.get("Rit")
            if not pd.isna(rv) and rv!="": rit_samenvatting[int(rv)]=row

    klant_ritten={}
    for klant,rit in zip(klanten,ritten):
        if pd.isna(rit): continue
        rit=int(rit)
        if klant not in klant_ritten: klant_ritten[klant]=set()
        klant_ritten[klant].add(rit)

    klant_rit_totalen={}
    for klant,rs in klant_ritten.items():
        klant_rit_totalen[klant]={"NL":0,"DE":0,"BE":0,"TOL":0,"KOSTEN":0}
        for rit in rs:
            if rit in rit_samenvatting:
                for k2,vk in [("NL","NL km"),("DE","DE km"),("BE","BE km"),("TOL","Tol km"),("KOSTEN","Kosten €")]:
                    klant_rit_totalen[klant][k2]+=rit_samenvatting[rit][vk]

    gezien_output=set(); output_results=[]
    for row in results:
        clean={k:v for k,v in row.items() if not k.startswith("_")}
        is_depot=row.get("Van")==BLEISWIJK_PLAATS and row.get("Naar")==BLEISWIJK_PLAATS
        if row.get("_is_totaal") or is_depot or not row.get("Opdracht nr"):
            output_results.append(clean); continue
        key=(row.get("Rit"),str(row.get("Opdracht nr") or ""))
        if key in gezien_output: continue
        gezien_output.add(key)
        clean["Gewicht (kg)"]=opdr_gew.get(key,"") or row.get("_gewicht") or ""
        output_results.append(clean)

    klant_rows=[]
    for klant,data in klant_rit_totalen.items():
        nr=len(klant_ritten[klant])
        klant_rows.append({"Klant":klant,"NL km":round(data["NL"],2),"DE km":round(data["DE"],2),
            "BE km":round(data["BE"],2),"Tol km":round(data["TOL"],2),
            "Kosten €":round(data["KOSTEN"],2),"Gem km per rit":round(data["TOL"]/nr if nr else 0,2)})

    save_cache(route_cache, ROUTE_CACHE_FILE)
    save_cache(geocode_cache, GEOCODE_CACHE_FILE)
    voortgang_bar.progress(1.0, text="✅ Klaar!")
    return output_results, klant_rows, rit_kaarten

# ─── UI ───────────────────────────────────────────────────────────────────────
st.title("🚛 VWH Tool — Kamps Transport")

with st.sidebar:
    st.header("⚙️ Instellingen")
    api_key     = st.text_input("ORS API Key", value="eyJvcmciOiI1YjNjZTM1OTc4NTExMTAwMDFjZjYyNDgiLCJpZCI6IjA2NjNjOTk3NGNiYjRmZjQ5YThhMjFhZWFmZjgyNWFmIiwiaCI6Im11cm11cjY0In0=", type="password")
    tarief      = st.number_input("Tarief per km (€)", value=0.23, step=0.01, format="%.3f")
    route_sleep = st.slider("Wachttijd ORS (sec)", 0.1, 3.0, 0.5, 0.1)
    st.divider()
    st.caption("Kolom G in Excel = Landcode (NL/DE/BE, optioneel)")

tab_upload, tab_resultaat, tab_kaart, tab_manueel = st.tabs([
    "📂 Upload & Berekenen", "📊 Resultaten", "🗺️ Kaart", "📍 Handmatige route"
])

# ── Tab 1: Upload & Berekenen ─────────────────────────────────────────────────
with tab_upload:
    uploaded = st.file_uploader("Upload je Excel-invoerbestand", type=["xlsx"])
    if uploaded:
        df_preview = pd.read_excel(uploaded)
        st.dataframe(df_preview.head(10), use_container_width=True)
        st.caption(f"{len(df_preview)} rijen geladen")

        if st.button("▶️ Start berekening", type="primary"):
            with st.spinner("Resources laden..."):
                tolwegen, tol_sindex, nl_geom, de_geom, be_geom, postcode_df, de_df, be_df = load_resources()
            geocode_cache = load_cache(GEOCODE_CACHE_FILE)
            route_cache   = load_cache(ROUTE_CACHE_FILE)
            voortgang_bar = st.progress(0, text="Starten...")
            status_tekst  = st.empty()
            uploaded.seek(0)
            df_input = pd.read_excel(uploaded)
            output_results, klant_rows, rit_kaarten = bereken(
                df_input, api_key, tarief, route_sleep,
                geocode_cache, route_cache,
                tolwegen, tol_sindex, nl_geom, de_geom, be_geom,
                postcode_df, de_df, be_df, voortgang_bar, status_tekst)
            st.session_state["output_results"] = output_results
            st.session_state["klant_rows"]     = klant_rows
            st.session_state["rit_kaarten"]    = rit_kaarten
            status_tekst.success("✅ Klaar! Ga naar Resultaten of Kaart.")

# ── Tab 2: Resultaten ─────────────────────────────────────────────────────────
with tab_resultaat:
    if "output_results" in st.session_state:
        output_results = st.session_state["output_results"]
        klant_rows     = st.session_state["klant_rows"]
        df_out = pd.DataFrame(output_results)
        st.dataframe(df_out, use_container_width=True, height=500)
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("Klant totalen")
            st.dataframe(pd.DataFrame(klant_rows), use_container_width=True)
        with col2:
            st.subheader("Samenvatting")
            totalen = [r for r in output_results if r.get("Van")=="TOTAAL"]
            if totalen:
                totaal_km  = sum(r.get("NL km",0)+r.get("DE km",0)+r.get("BE km",0) for r in totalen)
                totaal_tol = sum(r.get("Tol km",0) for r in totalen)
                st.metric("Totaal km", f"{totaal_km:,.0f} km")
                st.metric("Totaal tol km", f"{totaal_tol:,.0f} km")
                st.metric("% tol", f"{(totaal_tol/totaal_km*100 if totaal_km else 0):.1f}%")
                st.metric("Totale kosten", f"€ {sum(r.get('Kosten €',0) for r in totalen):,.2f}")
                st.metric("Aantal ritten", len(totalen))
        excel_buf = maak_excel(output_results, klant_rows)
        st.download_button("⬇️ Download Excel", data=excel_buf,
                           file_name="vwh_resultaat.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           type="primary")
    else:
        st.info("Voer eerst een berekening uit via het tabblad Upload & Berekenen.")

# ── Tab 3: Kaart ──────────────────────────────────────────────────────────────
with tab_kaart:
    if "rit_kaarten" in st.session_state:
        rit_kaarten = st.session_state["rit_kaarten"]
        ritten_lijst = sorted([r for r in rit_kaarten.keys() if rit_kaarten[r]])
        if ritten_lijst:
            gekozen_rit = st.selectbox("Kies een rit", ritten_lijst,
                                       format_func=lambda x: f"Rit {int(x)}")
            m = maak_kaart(f"Rit {int(gekozen_rit)}", rit_kaarten[gekozen_rit])
            if m: st_folium(m, width=None, height=550, returned_objects=[])
        else:
            st.warning("Geen kaartdata beschikbaar.")
    else:
        st.info("Voer eerst een berekening uit via het tabblad Upload & Berekenen.")

# ── Tab 4: Handmatige route ───────────────────────────────────────────────────
with tab_manueel:
    st.markdown("Voer stops in met plaats, postcode, landcode, opdrachtnummer en gewicht.")

    if "man_stops" not in st.session_state:
        st.session_state["man_stops"] = [
            {"plaats": "Bleiswijk", "postcode": "2665MZ", "land": "NL", "opdrachtnr": "", "gewicht": 0},
            {"plaats": "", "postcode": "", "land": "NL", "opdrachtnr": "", "gewicht": 0},
        ]

    stops = st.session_state["man_stops"]
    for i, stop in enumerate(stops):
        c1,c2,c3,c4,c5,c6 = st.columns([3,2,1,2,2,1])
        with c1: stops[i]["plaats"]     = st.text_input(f"Plaats {i+1}", value=stop["plaats"], key=f"p_{i}")
        with c2: stops[i]["postcode"]   = st.text_input(f"Postcode {i+1}", value=stop["postcode"], key=f"pc_{i}")
        with c3: stops[i]["land"]       = st.selectbox(f"Land {i+1}", ["NL","DE","BE",""], index=["NL","DE","BE",""].index(stop.get("land","NL")), key=f"l_{i}")
        with c4: stops[i]["opdrachtnr"] = st.text_input(f"Opdracht {i+1}", value=stop.get("opdrachtnr",""), key=f"o_{i}")
        with c5: stops[i]["gewicht"]    = st.number_input(f"Gewicht kg {i+1}", value=float(stop.get("gewicht",0)), min_value=0.0, step=100.0, key=f"g_{i}")
        with c6:
            st.write(""); st.write("")
            if i > 0 and st.button("🗑️", key=f"del_{i}"):
                st.session_state["man_stops"].pop(i); st.rerun()

    col_add, col_calc = st.columns([1,3])
    with col_add:
        if st.button("➕ Stop toevoegen"):
            st.session_state["man_stops"].append({"plaats":"","postcode":"","land":"NL","opdrachtnr":"","gewicht":0})
            st.rerun()
    with col_calc:
        bereken_btn = st.button("🗺️ Bereken route + COFRET", type="primary")

    if bereken_btn:
        geldig = [s for s in stops if s["plaats"].strip()]
        if len(geldig) < 2:
            st.error("Vul minimaal 2 stops in.")
        else:
            with st.spinner("Resources laden..."):
                tolwegen, tol_sindex, nl_geom, de_geom, be_geom, postcode_df, de_df, be_df = load_resources()
            geocode_cache = load_cache(GEOCODE_CACHE_FILE)
            route_cache   = load_cache(ROUTE_CACHE_FILE)
            depot_stop, depot_coords = geldig[0], None
            depot_coords = geocode(depot_stop["plaats"], depot_stop["postcode"], api_key,
                                   geocode_cache, postcode_df, de_df, be_df, depot_stop["land"] or None)
            alle_segmenten=[]; resultaat_rijen=[]; directe_routes={}
            tol_totaal=km_totaal=kosten_totaal=0
            prog = st.progress(0, text="Routes berekenen...")
            n = len(geldig)-1
            coords_lijst=[]
            for s in geldig:
                c = geocode(s["plaats"],s["postcode"],api_key,geocode_cache,postcode_df,de_df,be_df,s["land"] or None)
                coords_lijst.append((s,c))

            for i in range(n):
                stop_van, start = coords_lijst[i]
                stop_naar, end  = coords_lijst[i+1]
                prog.progress((i+1)/n, text=f"{stop_van['plaats']} → {stop_naar['plaats']}")
                if not start or not end:
                    st.warning(f"Kon {stop_van['plaats']} of {stop_naar['plaats']} niet geocoden.")
                    continue
                route = get_route(start,end,route_cache,api_key,tolwegen,tol_sindex,nl_geom,de_geom,be_geom,route_sleep)
                route_direct = get_route(depot_coords,end,route_cache,api_key,tolwegen,tol_sindex,nl_geom,de_geom,be_geom,route_sleep) if depot_coords else None
                if route:
                    nl,de,be,tol,coords,segmenten=route
                    alle_segmenten.extend(segmenten)
                    km=nl+de+be; kosten=tol*tarief
                    tol_totaal+=tol; km_totaal+=km; kosten_totaal+=kosten
                    direct_tol=direct_km=direct_kosten=0
                    if route_direct:
                        nl_d,de_d,be_d,direct_tol,_,_=route_direct
                        direct_km=nl_d+de_d+be_d; direct_kosten=direct_tol*tarief
                    opdr=stop_naar.get("opdrachtnr","").strip() or f"stop_{i+1}"
                    gew=float(stop_naar.get("gewicht",0) or 0)
                    directe_routes[opdr]={"gewicht":gew,"direct_tol":direct_tol,"direct_km":direct_km,"direct_kosten":direct_kosten}
                    resultaat_rijen.append({"Van":stop_van["plaats"],"Naar":stop_naar["plaats"],
                        "Opdracht nr":opdr,"Gewicht (kg)":gew,
                        "NL km":round(nl,2),"DE km":round(de,2),"BE km":round(be,2),
                        "Tol km":round(tol,2),"Kosten rit €":round(kosten,2),
                        "Km direct":round(direct_km,2),"Tol km direct":round(direct_tol,2),
                        "Kosten direct €":round(direct_kosten,2)})
            prog.empty()
            save_cache(route_cache, ROUTE_CACHE_FILE)
            save_cache(geocode_cache, GEOCODE_CACHE_FILE)

            # COFRET
            opdr_vwh={}
            for opdr,d in directe_routes.items():
                bezet=d["gewicht"]/MAX_GEWICHT if d["gewicht"]>0 else 0
                opdr_vwh[opdr]=d["direct_kosten"]*bezet
            vwh_som=sum(opdr_vwh.values())
            direct_km_som=sum(d["direct_km"] for d in directe_routes.values())
            for row in resultaat_rijen:
                opdr=row["Opdracht nr"]
                vwh=opdr_vwh.get(opdr,0)
                row["Kosten toegerekend €"]=round((vwh/vwh_som)*kosten_totaal,2) if vwh and vwh_som else 0
                dk=directe_routes.get(opdr,{}).get("direct_km",0)
                row["Km toegerekend"]=round((dk/direct_km_som)*km_totaal,2) if dk and direct_km_som else 0

            # Metrics
            c1,c2,c3,c4=st.columns(4)
            c1.metric("Totaal km",f"{km_totaal:.2f} km")
            c2.metric("Tol km",f"{tol_totaal:.2f} km")
            c3.metric("% tol",f"{(tol_totaal/km_totaal*100 if km_totaal else 0):.1f}%")
            c4.metric("Kosten tol",f"€ {kosten_totaal:.2f}")

            st.subheader("Resultaat per stop")
            st.dataframe(pd.DataFrame(resultaat_rijen), use_container_width=True)

            if alle_segmenten:
                st.subheader("Kaart")
                m = maak_kaart("Handmatige route", alle_segmenten)
                if m: st_folium(m, width=None, height=550, returned_objects=[])
